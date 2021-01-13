import requests
import json
import openpyxl
import datetime
import time
import sys
import os


class HeroRec():
    def __init__(self):
        try:
            with open(os.path.dirname(os.path.abspath(sys.argv[0])) + "/config.json", mode="r") as jf:
                df = json.load(jf)
                self.levellist = df["levellist"]
                self.levelidlist = df["levelidlist"]
                self.weaponid = df["weaponid"]
                self.weaponlist = df["weaponlist"]
                self.weaponiddict = df["weaponiddict"]
                self.invalidrunlist = df["invalidrunlist"]  # not used
        except FileNotFoundError:
            sys.exit("config.json does not found.")

    def opensheet(self):
        try:
            wbpath = os.path.dirname(os.path.abspath(sys.argv[0])) + "/HeromodeILrecords.xlsx"
            self.wb = openpyxl.load_workbook(wbpath)
            self.ws = self.wb.worksheets[0]
        except FileNotFoundError:  # file is not exist
            self.wb = openpyxl.Workbook()  # create new file
            print("HeromodeILrecords.xlsx does not found.")
            print("set up new sheet.")
            self.ws = self.wb.worksheets[0]
            self.sheetsetup()

    def cell(self, row, column):
        return self.ws.cell(row+2, column+2)  # set (2, 2) as the origin

    def sheetsetup(self):
        # level
        n = 0
        level = 1
        for i in range(5):
            for _ in range([3, 6, 6, 6, 6][i]):
                self.cell(n, -1).value = f"{level:0>2}"
                self.cell(n, 11).value = f"{level:0>2}"
                n += 1
                level += 1
            self.cell(n, -1).value = f"Boss{i+1}"
            self.cell(n, 11).value = f"Boss{i+1}"
            n += 1
        self.cell(n, -1).value = "sum"
        self.cell(n, 11).value = "sum"

        # weapon
        for m in range(9):
            self.cell(-1, m).value = self.weaponlist[m]
        self.cell(-1, 9).value = "sum"
        self.cell(-1, 12).value = "WR"
        self.cell(-1, 13).value = "weapon"

        # format
        for n in range(32+1):
            for m in range(9+1):
                self.cell(n, m).number_format = "[mm]:ss"
            self.cell(n, 12).number_format = "[mm]:ss"

    def getrec(self, levelnum):
        levelurl = f"https://www.speedrun.com/api/v1/runs?status=verified&max=200&level={self.levelidlist[levelnum]}"  # first page url
        leveldata = []  # runデータは一度に200個までしか入手できないので200個ずつ取得してleveldataにまとめて格納しておく
        while True:
            response = requests.get(levelurl)
            response.raise_for_status()
            time.sleep(5)
            levelinfo = response.json()
            leveldata += levelinfo["data"]
            # search for next page
            levellinks = levelinfo["pagination"]["links"]
            for link in levellinks:
                if link["rel"] == "next":
                    levelurl = link["uri"]  # get next page url
                    break
            else: # next page is not exist
                break

        # assort by weapon
        weaponrec = [
            [
                rundata["times"]["primary_t"]
                for rundata in leveldata
                if self.weaponiddict[rundata["values"][self.weaponid]] == n
                and rundata["id"] not in self.invalidrunlist
            ]
            for n in range(9)
        ]
        weaponrec = [
            min(weaponrec[n])
            if weaponrec[n]
            else False
            for n in range(9)
        ]
        # 次と同値
        """
        weaponrec = [False]*9
        # check all run data
        for rundata in leveldata:
            if not rundata["id"] in self.invalidrunlist:
                runweaponid = rundata["values"][self.weaponid]
                weaponnum = self.weaponiddict[runweaponid]
                runrec = rundata["times"]["primary_t"]
                currentrec = weaponrec[weaponnum]
                if not currentrec or currentrec > runrec:
                    weaponrec[weaponnum] = runrec  # renew WR
        """

        return weaponrec

    def main(self):
        self.opensheet()
        WRlist = [None]*32
        print("get all records from Speedrun.com.")
        for n in range(32):
            WRlist[n] = self.getrec(n)
            print(f"{self.levellist[n]} done.")

        def inttotime(t):
            s = datetime.time(
                t//3600, # hour
                t%3600//60, # minute
                t%60, # second
            )
            return s

        # find new record
        for n in range(32):
            for m in range(9):
                if WRlist[n][m]:
                    if self.cell(n, m).value:
                        if inttotime(WRlist[n][m]) < self.cell(n, m).value:
                            print(f"New record! {self.levellist[n]} {self.weaponlist[m]}, {WRlist[n][m]//60:0>2}:{WRlist[n][m]%60:0>2}")
                    else:
                        print(f"New record! {self.levellist[n]} {self.weaponlist[m]}, {WRlist[n][m]//60:0>2}:{WRlist[n][m]%60:0>2}")

        # データの整形
        awWRlist = []  # any weapon WR
        for n in range(32):
            tmp = [WRlist[n][m] for m in range(9) if WRlist[n][m]] # exclude False
            if tmp:
                awWRlist.append(min(tmp))
            else:
                awWRlist.append(False)
        if all(awWRlist):
            awWRlist.append(sum(awWRlist))
        else:
            awWRlist.append(False)
<<<<<<< HEAD
        # sum of all weapon records
=======
>>>>>>> 289440d7cab9547f188fe39a87e1bff4bbfe897d
        for n in range(32):
            if all(WRlist[n]):
                WRlist[n].append(sum(WRlist[n]))
            else:
                WRlist[n].append(False)
        WRlist.append([])
        # sum of all level records
        for m in range(9+1):
            tmp = [WRlist[n][m] for n in range(32)]
            if all(tmp):
                WRlist[32].append(sum(tmp))
            else:
                WRlist[32].append(False)

        # renew all records in sheet
        for n in range(32+1):
            for m in range(9+1):
                if WRlist[n][m]:
                    self.cell(n, m).value = inttotime(WRlist[n][m])
                else:
                    self.cell(n, m).value = ""
            if awWRlist[n]:
                self.cell(n, 12).value = inttotime(awWRlist[n])
            else:
                self.cell(n, 12).value = ""
        for n in range(32):
            self.cell(n, 13).value = self.weaponlist[WRlist[n].index(awWRlist[n])]

        self.cell(-1, -1).value = datetime.datetime.now().strftime("%Y/%m/%d")
        self.wb.save("HeromodeILrecords.xlsx")


if __name__ == "__main__":
    a = HeroRec()
    a.main()
